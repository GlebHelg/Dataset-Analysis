import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
import csv
import sys

# Global #####################################################
header = [     "NRF",    "BD_11",  "ONN_7",   "HR_3", "AHL_2"]
matrix = []
##############################################################

firstIt=True
pkgCount=0
prePkg=0

#Global verdi for aa kalkulere pakkepris
pkgPrice = [0]*4

#Takes NRF, returns following list: [found_idx, status]
#Status codes: 
# -2 = !found nrf, and in end of list     : matrix.append
# -1 = !found nrf, and not in end of list : matrix.insert
#  0 =  found nrf, insert new value in matrix_row
def binarySearch(nrf):
    global matrix
    first    = 0
    last     = len(matrix)-1
    found    = False
    midpoint = 0

    while first<=last and not found:
        midpoint = (first + last)/2
        if matrix[midpoint][0] == nrf:
            found = True
        else:
            if nrf < matrix[midpoint][0]:
                last = midpoint-1
            else:
                first = midpoint+1
    
    return midpoint

#END: binarySearch()

def rowExtractor(nrf):
    print nrf
    midpoint = binarySearch(nrf)

    row = matrix[midpoint][1:]

    return row

#Takes input, and finds correct way to insert into matrix, then inserts
def fillMatrix(nrf, lev, price, l):
    global matrix

    newInsertion = "Not a list"
    matrIdx = 0

    if lev == 2:
        #AHL
        newInsertion = [nrf,0,0,0,price]
        matrIdx = 4
    elif lev == 3:
        #HR
        newInsertion = [nrf,0,0,price,0]
        matrIdx = 3
    elif lev == 7:
        #ONN
        newInsertion = [nrf,0,price,0,0]
        matrIdx = 2
    elif lev == 11:
        #BD
        newInsertion = [nrf,price,0,0,0]
        matrIdx = 1
    else:
        print "Unknown grossist." 
        print "Error happened at loop-iteration: ", l
        print "Nothing inserted this loop"
        return
    
    if len(matrix) == 0:
        matrix.append(newInsertion)
        return

    midpoint = binarySearch(nrf)

    if matrix[midpoint][0] == nrf:
        #Insert price in existing row
        matrix[midpoint][matrIdx] = price
    
    elif matrix[midpoint][0] > nrf:
        #Insert new row
        matrix.insert(midpoint, newInsertion)

    elif matrix[midpoint][0] < nrf:
        #Insert new row
        if midpoint != len(matrix)-1:
            #After
            matrix.insert(midpoint+1, newInsertion)
        else:
            #Append
            matrix.append(newInsertion)

#END: fillMatrix()

def xmlParser(loopList, prodList):

    for xml_file in loopList:

        #Reading from file and parsing XML
        tree = ET.parse(xml_file)
        root = tree.getroot()

        datagrunnlag = root
        varekataloger = datagrunnlag[4]

        for varekatalog in varekataloger.iter('VAREKATALOG'):

            leverId  = int(varekatalog.find('LEVERID').text)

            if leverId == 1337:
                continue

            for produkt in varekatalog.iter('PRODUKT'):
                #[nrf, lev, price]
                prodList.append([produkt[0].text, leverId, float(produkt[12].text)])
        
        return prodList
#END: xmlParser()


#Faar pris basert paa antall som maa bestilles ved kjoop, returnerer liste paa storrelse 4
def priceByQuant(priceList, active, row):
    cell = "{}{}".format('L', row)
    for i in range(0,4):
        priceList[i]=priceList[i]*active[cell].value
    return priceList

#Finner laveste pris paa modul-nivaa, returnerer index paa denne
def findLowest(priceList):
    for i in range(0,4):
        int(priceList[i])

    low=10000000
    index=0
    for i in range(0,4):
        if low>priceList[i] and priceList[i]!=0:
            low=priceList[i]
            index = i
    print index
    print priceList
    return index

# Finner summen av varer i en pakke, og returnerer denne
def bestPkgSum():
    global pkgPrice
    if sum(pkgPrice)==0:
        return 0
    return min(i for i in pkgPrice if i > 0)

# Finner leverandor med lavest pakkepris
def bestPkgLeve():
    global pkgPrice
    if sum(pkgPrice)==0:
        return "PKG MIA"
    value = min(i for i in pkgPrice if i > 0)
    index = pkgPrice.index(value)
    if index == 0:
        return "BD"
    elif index == 1:
        return "ONN"
    elif index == 2:
        return "HEID"
    else:
        return "AHL"
#Hovedfunksjon for aa kalkulere prisen paa pakker, samt putter dette i excel ark
def packageCalculation(priceList, row, active, currentPkgNrf):
    global pkgPrice, pkgCount, prePkg, firstIt
    if currentPkgNrf != None and currentPkgNrf!=prePkg and firstIt == True:
        pkgCount=0
        firstIt=False
        pkgCount+=1
        prePkg=currentPkgNrf
        for i in range(0,4):
            if priceList[i] != 0:
                pkgPrice[i]+=priceList[i]

    elif currentPkgNrf != None and currentPkgNrf!=prePkg:
        prePkg=currentPkgNrf
        pkgCount+=1
        bestPrice=bestPkgSum()
        bestLeve=bestPkgLeve()
        pkgRow=row-pkgCount
        pkgCount=0
        cell_bestLeve = "{}{}".format('F',pkgRow)
        cell_bestPrice = "{}{}".format('G',pkgRow)
        print cell_bestLeve
        active[cell_bestLeve]=bestLeve
        active[cell_bestPrice]=bestPrice
        pkgPrice=[0,0,0,0]
        for i in range(0,4):
            if priceList[i] != 0:
                pkgPrice[i]+=priceList[i]
    elif row == active.max_row+1:
        pkgCount+=1
        bestPrice=bestPkgSum()
        bestLeve=bestPkgLeve()
        pkgRow=row-pkgCount
        pkgCount=0
        cell_bestLeve = "{}{}".format('F',pkgRow)
        cell_bestPrice = "{}{}".format('G',pkgRow)
        active[cell_bestLeve]=bestLeve
        active[cell_bestPrice]=bestPrice
        
    else:
        pkgCount+=1
        for i in range(0,4):
            if priceList[i] != 0:
                if pkgPrice[i]!=0:
                    pkgPrice[i]+=priceList[i]
            else:
                pkgPrice[i]=0

#returnerer nrf til current pkg
def currentPkgNrf(row,active):
    currentPkg = "{}{}".format('B',row)
    return active[currentPkg].value

#Putter modulpriser i excel
def putInExcel(priceList, row, active):
    i=0
    for cell in "NOPQ":
        cell_name = "{}{}".format(cell,row)
        active[cell_name]=priceList[i]
        i+=1
    return True
#Et misslykket forsok av farging i excel
def colorBestPrice(priceList, row, active):
    indexBest = findLowest(priceList)
    i=0
    for cell in "NOPQ":
        if i==indexBest:
            cell_name= "{}{}".format(cell,row)
            active[cell_name].fill = PatternFill(fgColor='FFFFFF',bgColor='6EF4A0',fill_type='solid')
            break
        else:
            i+=1
    return True





#Hoved funksjon for arbeid i excel
def excelWork(inFile, destFile):
    global firstIt
    wb = load_workbook(inFile)
    for i in range(2,3):
        print("currently on page "+ str(i))
        sheet = wb.get_sheet_names()[i]
        active = wb.get_sheet_by_name(sheet)
        if i==2 or i==6 or i== 7 or i==8:
            startAt=5
        else:
            startAt=4
        firstIt=True
        pkgCount=0
        for row in range(startAt, active.max_row + 1):
            for cell in "H":
                cell_name = "{}{}".format(cell,row)
                priceList=rowExtractor(active[cell_name].value)
                newPriceList= priceByQuant(priceList,active ,row)
                currentPkgNrfval = currentPkgNrf(row,active)
                packageCalculation(newPriceList, row, active, currentPkgNrfval)
                putInExcel(newPriceList, row, active)
                #colorBestPrice(priceList, row, active)
        
            
    if destFile == '':
        wb.save("test2excel.xlsx")
    else:
        wb.save(destFile)

#Main som tar haand om linjeargumenter og startup av hovedfunksjoner i programmet
def main(argv):
    global header
    global matrix

    loopList = ['Datagrunnlag.xml']
    #loopList = ['bd.xml']

    prodList = []
    prodList = xmlParser(loopList, prodList)
    print "prodList ready. length: ", len(prodList)

    for l, prod in enumerate(prodList):
        if l % 10000 == 0:
            print l
        fillMatrix(prod[0], prod[1], prod[2], l)

    inFile = argv[2]
    destFile = ''

    excelWork(inFile,destFile)
    sys.exit
    try:
      opts, args = getopt.getopt(argv,"hi:o:",["ifile=","ofile="])
    except getopt.GetoptError: 
      print 'xmlToExcel.py -i <inputfile> -o <outputfile>'
      sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            println("xmlToExcel.py -i <inputfile> -o <outputfile>")
            println("This code is made to manipulate an excelfile, with the module standard made by Rorkjop AS.")
            sys.exit()
        elif opt in ("-i", "--ifile"):
            inputfile = arg
        elif opt in ("-o", "--ofile"):
            outputfile = arg
        
    print 'Input file is "', inputfile
    print 'Output file is "', outputfile

if __name__ == "__main__":
   main(sys.argv)
