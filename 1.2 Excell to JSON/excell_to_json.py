import copy
import math
import openpyxl
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import json
import argparse

"""
#Command Line Argument Parser
parser = argparse.ArgumentParser(description = "Specify IP, DB, PORT and USER")
parser.add_argument(  '--ip', default='localhost')
parser.add_argument('--port', default='5432'     )
parser.add_argument(  '--db', default='rk'      )
parser.add_argument(   '--u', default='postgres' )
"""

#    MUST FIX ARGPARSE FOR INPUT FILE
#Opening Excel file for globale
excell_file = "nye_pakker_btw.xlsx"
wb = load_workbook(excell_file)

sheet = wb.get_sheet_names()[2]
active = wb.get_sheet_by_name(sheet)

listOfSeries = []


def buildLevObj(lev_navn, pris_liste):
    lo_tp = 0
    for pris in pris_liste:
        lo_tp += pris
    lev_obj = {"lev_navn":lev_navn, "pris_liste":pris_liste, "tot_pris":lo_tp}
    return lev_obj
#END buildLevObj()

# EXCELL
def findPkgLenght_Nrf(rows):
    global active
    pkgLength = ["",0]
    for i in range(rows, rows+10):
        pre_cell = "{}{}".format("B",rows)
        next_cell = "{}{}".format("B",i)
        if active[pre_cell].value != active[next_cell].value or active[next_cell]==None:
            pkgLength[0] = str(active[pre_cell].value)
            return pkgLength
        else:
            pkgLength[1]+=1
#END findPkgLength()
    
def getExcellVals(pack_nrf, row_count, row_at):
    global active
    bigList = []
    bd_list = []
    on_list = []
    hr_list = []
    ahl_list = []
    nrf_list = [pack_nrf]
    for i in range(row_at, row_at+row_count):
        for x in "HNOPQ":
            cell_name = "{}{}".format(x,i)
            if active[cell_name].value != None:
                if x == "H":
                    nrf_list.append(str(active[cell_name].value))
                elif x == "N":
                    bd_list.append(float(active[cell_name].value))
                elif x == "O":
                    on_list.append(float(active[cell_name].value))
                elif x == "P":
                    hr_list.append(float(active[cell_name].value))
                elif x == "Q":
                    ahl_list.append(float(active[cell_name].value))
                else:
                    print "Someting went wrong in getExcellVals"
    bigList.append(nrf_list)
    bigList.append(bd_list)
    bigList.append(on_list)
    bigList.append(hr_list)
    bigList.append(ahl_list)
    return bigList
#END getExcellVals()

def getBestLevAndPrice(row_at, row_count):
    global active
    best_list = ["Not Complete pkg", 0]
    for i in "FG":
        cell_name = "{}{}".format(i,row_at+row_count)
        if i == "F" and active[cell_name].value != "PKG MIA":
            best_list[0] = active[cell_name].value
        elif i == "G":
            best_list[1] = active[cell_name].value

    return best_list
#END getBestLevAndPrice()

def getProdPackMatrix(pack_nrf, row_count, row_at):
    #Extract values from excell for one pack
    excellVals = getExcellVals(pack_nrf, row_count, row_at)
    nrf_list = excellVals[0][1:]
    bd_price_list = excellVals[1]
    onn_price_list = excellVals[2]
    hr_price_list = excellVals[3]
    ahl_price_list = excellVals[4]
    best_lev_price = getBestLevAndPrice(row_at, row_count)
    billigst_lev = best_lev_price[0]
    billigst_pris = best_lev_price[1]

    excell_prod_pack_matrix = [nrf_list, bd_price_list, onn_price_list, hr_price_list, ahl_price_list, billigst_lev, billigst_pris]
    return excell_prod_pack_matrix
#END getProdPackMatrix

# Builds one JSON product pack
def buildProdPack(pack_nrf, row_count, row_at):
    prod_pack = {"pack_nrf":pack_nrf, "modul_nrf":[], "lev_obj_list":[], "billigst_lev":"", "billigst_pris":999}
    #Bygger leverandor_obj_liste
    prod_pack_matrix = getProdPackMatrix(pack_nrf, row_count, row_at)
    prod_pack['modul_nrf'] = prod_pack_matrix[0]

    lev_obj_list = []
    lev_liste = ["BD","ONN","HR","AHL"]

    for idx, lev_navn in enumerate(lev_liste):
        if idx == 0:
            lev_pris_liste = prod_pack_matrix[1]
        elif idx == 1:
            lev_pris_liste = prod_pack_matrix[2]
        elif idx == 2:
            lev_pris_liste = prod_pack_matrix[3]
        elif idx == 3:
            lev_pris_liste = prod_pack_matrix[4]
        
        lev_obj = buildLevObj(lev_navn, lev_pris_liste)
        lev_obj_list.append(lev_obj)

    prod_pack['lev_obj_list'] = lev_obj_list
    prod_pack['billigst_lev'] = prod_pack_matrix[5]
    prod_pack['billigst_pris'] = prod_pack_matrix[6]

    return prod_pack
#END buildProdPack()

# Builds packages pr excell sheet
def buildPakker(x):
    global active, sheet, wb, listOfSeries
    pakker = []
    i = 5
    pack_rows=0

    while i < active.max_row:
        count_pack_rows = findPkgLenght_Nrf(i)
        pack_nrf  = count_pack_rows[0] 
        pack_rows = count_pack_rows[1]
        prod_pack = buildProdPack(pack_nrf, pack_rows, i)
        #Here we must insert cheapest combo and evaluate conflict

        pakker.append(prod_pack)
        i+=pack_rows
    
    return pakker
#END buildPakker()

def checkIfSub(preNrfs, currentNrfs):
    countEquals = 0
    if len(preNrfs)!=len(currentNrfs):
        return False
    for x in range(len(currentNrfs)):
        if preNrfs[x]==currentNrfs[x]:
            countEquals+=1
    precentageCalc = int(math.ceil(len(preNrfs)/2.0))
    if countEquals>=precentageCalc:
        return True
    else:
        return False



##Making a list of series
def serieByNrf(data, i):
    subSeries = True
    totalSeries = True
    counter = 0
    seriePkg = []
    listOfNrf = []
    listOfNrf.append(data[0]["modul_nrf"])
    
    for i in range(len(data)):
        if i==0:
             seriePkg.append(data[i])

        elif checkIfSub(data[i-1]["modul_nrf"],data[i]["modul_nrf"]):
             seriePkg.append(data[i])
        
        else:
            return seriePkg

    return seriePkg
#END serieByNrf()

##makes serie json and returns this so referances dont arent messing with values
def makeSubSerie(i, packages):
   serie = {"sub_serie_id":i, "pakker":packages}
   return serie 
#END makeSubSerie()
"""
def findSuppSerieCombo(sub_serie):
    cheapest_supplier_in_serie_combo = []
    for pack in sub_serie:
        cheapest_supplier_in_serie_combo.append(pack["billigst_lev"])
    sub_serie["billigste_leverandorer_i_serie"] = []
    sub_serie["billigste_leverandorer_i_serie"] = cheapest_supplier_in_serie_combo
"""
def makeSubSeriesList(data):
    i = 0
    subSeriesList = []
    for i in range(len(data)):
        subSeriesList.append(makeSubSerie(i, serieByNrf(data, i)))

    return subSeriesList
#END makeSerieList()

"""
# Find cheapest supplier combo on module level
def findCheapestModCombo(package):
    billigste_lev_combo = []
    billigste_lev_combo_pris = 0

    lev_obj_list = package['lev_obj_list']

    length_of_lev_obj_list = len(lev_obj_list)
    length_of_pricelist = len(lev_obj_list[0]['pris_liste'])

    lev_obj_list_idx_counter = 0
    for i in range(0, length_of_pricelist):
        billigste_lev = "none"
        current_cheapest_price = 99999999
        for j in range(0, length_of_lev_obj_list):
            if lev_obj_list[j]['pris_liste'][i] != 0 and lev_obj_list[j]['pris_liste'][i] < current_cheapest_price:
                current_cheapest_price = lev_obj_list[j]['pris_liste'][i]
                billigste_lev = lev_obj_list[j]['lev_navn']
        billigste_lev_combo_pris += current_cheapest_price
        billigste_lev_combo.append(billigste_lev)

    return [billigste_lev_combo, billigste_lev_combo_pris]
#END findChepestModCombo()

def cheapestModComboLoop(pack_list):
    for package in pack_list:
        theCombo = findCheapestModCombo(package)
        package['billigste_lev_combo'] = theCombo[0]
        package['billigste_lev_combo_pris'] = theCombo[1]
#END cheapestModComboLoop()
"""

def makeSeriesList():
    series_list = []
    i=0
    for x in range(2,10):
        print "Now on page:"+str(x)
        serie = {"serie_id":x-2,"sub_series_list":[]}

        packs_in_sheet = buildPakker(x)
        # Finds cheapest combo on module level, and calculates the price
        #cheapestModComboLoop(packs_in_sheet)
        # 1. Make sub_series-list of these packs
        serie["sub_series_list"] = makeSubSeriesList(packs_in_sheet)
        # 2. append sub_series-list to series
        series_list.append(serie)
        #print sub_series
    return series_list

root = {"series":[]}
root["series"] = makeSeriesList()
f = open('test2.json','wb')
f.write(json.dumps(root,indent=4))